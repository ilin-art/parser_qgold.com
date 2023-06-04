import argparse
import requests
import time
import os
import openpyxl
import concurrent.futures


def fetch(url: str, params: dict[str, any]) -> requests.Response:
    """
    Выполняет HTTP-запрос на указанный URL с заданными параметрами.
    """
    headers = params['headers']
    body = params['body']
    if params['method'] == 'POST':
        return requests.post(url, headers=headers, data=body)
    else:
        return requests.get(url, headers=headers)

def parse_products(data: list[dict[str, any]]) -> list[str]:
    """
    Извлекает URL-адреса продуктов со страницы с товарами.
    """
    urls = []
    for product in data:
        url_description = product["URLDescription"]
        style = product["Style"]
        url = f"https://jewelers.services/productcore/api/pd/{url_description}/{style}"
        urls.append(url)
    return urls

def process_data(fetch: requests.Response) -> list[str]:
    """
    Обрабатывает данные о продуктах и сохраняет их в файл Excel.
    """
    result = fetch.json()
    indexed_products = result.get("IndexedProducts")
    if indexed_products:
        results = indexed_products.get("Results")
        if results:
            urls = parse_products(results)
            file_name = process_file_name(fetch.url)
            process_excel(file_name, fetch_urls(urls))
            return urls
        else:
            print("Нет данных о продуктах")
    else:
        print("Нет данных в IndexedProducts")

def process_file_name(url: str) -> str:
    """
    Генерирует имя файла на основе URL-адреса.
    """
    file_name = os.path.basename(url)
    file_name = file_name.replace("?", "").replace("/", "_")
    file_name = file_name.split("=")[0] + ".xlsx"
    return file_name

def fetch_urls(urls: list[str]) -> list[list[str]]:
    """
    Получает данные для каждого URL-адреса продукта.
    """
    data_list = []
    params = {
        "headers": {
            "accept": "application/json, text/plain, */*",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "authorization": "",
            "cache-control": "no-cache",
            "content-type": "application/json",
            "pragma": "no-cache",
            "sec-ch-ua": "\"Not.A/Brand\";v=\"8\", \"Chromium\";v=\"114\", \"Google Chrome\";v=\"114\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "cross-site",
            "trailerid": "04c24bde-feb7-4b27-8672-641991888d08",
            "Referer": "https://qgold.com/",
            "Referrer-Policy": "strict-origin-when-cross-origin"
        },
        "body": None,
        "method": "GET"
    }
    for url in urls:
        response = fetch(url, params)
        if response.status_code == 200:
            result = response.json()
            description = result.get('Product', {}).get('Description')
            sizes = result.get('Sizes')
            specifications = result.get('Specifications')
            images = result.get('Images')
            video = result.get('Video')
            availability = result.get('Product', {}).get('AvailabilityText')
            specification_values = []
            for spec in specifications:
                spec_name = spec.get('Specification')
                spec_value = spec.get('Value')
                specification_values.append(f"{spec_name}: {spec_value}")
            spec_string = "; ".join(specification_values)
            image_names = [image.get('FileName') for image in images]
            image_urls = [f"https://images.jewelers.services/qgrepo/{image_name}" for image_name in image_names]
            image_string = "; ".join(image_urls)
            video_filename = video.get('FileName') if video else "None"
            video_url = f"https://images.jewelers.services/0/Videos/{video_filename}" if video else "None"
            if sizes:
                size_price_list = [f"{size_data.get('Size')} - {size_data.get('MSRP')}" for size_data in sizes]
                size_price_string = "; ".join(size_price_list)
                data_list.append([description, size_price_string, spec_string, image_string, video_url, availability])
            else:
                msrp = result.get('Product', {}).get('MSRP')
                data_list.append([description, msrp, spec_string, image_string, video_url, availability])
        else:
            print(f"Ошибка при получении данных для URL: {url}")
    return data_list

def process_excel(file_name: str, data: list[list[str]]) -> None:
    """
    Создает или обновляет файл Excel с данными.
    """
    if os.path.exists(file_name):
        update_excel(file_name, data)
    else:
        create_excel(file_name, data)

def create_excel(file_name: str, data: list[list[str]]) -> None:
    """
    Создает новый файл Excel и сохраняет данные в него.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Name", "Size - Price", "Specifications", "Images", "Video", "Availability"])
    for row in data:
        worksheet.append(row)
    workbook.save(file_name)

def update_excel(file_name: str, data: list[list[str]]) -> None:
    """
    Обновляет существующий файл Excel с новыми данными или удаляет устаревшие данные.
    """
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active
    name_column = worksheet["A"][1:]
    name_index_dict = {}
    for i, cell in enumerate(name_column, start=2):
        name = cell.value
        if name:
            name_index_dict[name] = i
    delete_rows = []
    for name, row_index in name_index_dict.items():
        if name not in [row[0] for row in data]:
            delete_rows.append(row_index)
    for row_index in reversed(delete_rows):
        worksheet.delete_rows(row_index)
    for row in data:
        name = row[0]
        if name and name in name_index_dict:
            row_index = name_index_dict[name]
            for j, value in enumerate(row, start=1):
                cell_value = worksheet.cell(row=row_index, column=j).value
                if cell_value != value:
                    worksheet.cell(row=row_index, column=j).value = value
        else:
            empty_row_index = 1
            while worksheet.cell(row=empty_row_index, column=1).value:
                empty_row_index += 1
            for j, value in enumerate(row, start=1):
                worksheet.cell(row=empty_row_index, column=j).value = value
    workbook.save(file_name)

jewelry_rings_adjustable_fetch = fetch("https://jewelers.services/productcore/api/pl/Jewelry-Rings-Adjustable?v=Sun%20Jun%2004%202023%2012:23:26%20GMT+0300%20(%D0%9C%D0%BE%D1%81%D0%BA%D0%B2%D0%B0,%20%D1%81%D1%82%D0%B0%D0%BD%D0%B4%D0%B0%D1%80%D1%82%D0%BD%D0%BE%D0%B5%20%D0%B2%D1%80%D0%B5%D0%BC%D1%8F)", {
  "headers": {
    "accept": "application/json, text/plain, */*",
    "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "authorization": "",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "pragma": "no-cache",
    "sec-ch-ua": "\"Not.A/Brand\";v=\"8\", \"Chromium\";v=\"114\", \"Google Chrome\";v=\"114\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "cross-site",
    "trailerid": "04c24bde-feb7-4b27-8672-641991888d08",
    "Referer": "https://qgold.com/",
    "Referrer-Policy": "strict-origin-when-cross-origin"
  },
  "body": "{\"filters\":[{\"key\":\"ItemsPerPage\",\"value\":\"36\"}],\"page\":1,\"sortCode\":5,\"path\":\"Jewelry-Rings-Adjustable\"}",
  "method": "POST"
})

Jewelry_Rings_CZ_Rings_fetch = fetch("https://jewelers.services/productcore/api/pl/Jewelry-Rings-CZ-Rings?v=Sun%20Jun%2004%202023%2011:44:13%20GMT+0300%20(%D0%9C%D0%BE%D1%81%D0%BA%D0%B2%D0%B0,%20%D1%81%D1%82%D0%B0%D0%BD%D0%B4%D0%B0%D1%80%D1%82%D0%BD%D0%BE%D0%B5%20%D0%B2%D1%80%D0%B5%D0%BC%D1%8F)", {
  "headers": {
    "accept": "application/json, text/plain, */*",
    "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "authorization": "",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "pragma": "no-cache",
    "sec-ch-ua": "\"Not.A/Brand\";v=\"8\", \"Chromium\";v=\"114\", \"Google Chrome\";v=\"114\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "cross-site",
    "trailerid": "04c24bde-feb7-4b27-8672-641991888d08",
    "Referer": "https://qgold.com/",
    "Referrer-Policy": "strict-origin-when-cross-origin"
  },
  "body": "{\"filters\":[{\"key\":\"ItemsPerPage\",\"value\":\"36\"}],\"page\":1,\"sortCode\":5,\"path\":\"Jewelry-Rings-CZ-Rings\"}",
  "method": "POST"
})

# Определение аргументов командной строки
parser = argparse.ArgumentParser(description='Parser options')
parser.add_argument('--interval', type=int, default=900, help='Interval in seconds between parsing requests')
args = parser.parse_args()

# Цикличный запуск функций в параллельных потоках с заданной периодичностью
while True:
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        futures.append(executor.submit(process_data, jewelry_rings_adjustable_fetch))
        futures.append(executor.submit(process_data, Jewelry_Rings_CZ_Rings_fetch))
    time.sleep(args.interval)
